#importing library
import math

# other style of import
import math as m
from math import *
from math import pi, cos

# using math module variable
print("pi number: ", math.pi)

# using math module function
print("cos of 2*pi: ",math.cos(math.pi*2))

# import random module
import random

# use random module function
print("random integer: ", random.randint(0,2))

# Using library to work with excel file

import openpyxl as xl

# open workbook
wb = xl.load_workbook("example/example.xlsx")

# print sheet names
print("sheets name: ", wb.sheetnames)

# open sheet
sheet_1 = wb['Sheet1']

# read value from file
print("cell A1 value: ",sheet_1.cell(1,1).value)

# write value to cell
sheet_1.cell(4,1, "Yessy M")

# show name of tables in a sheet
print("tables' name: ",sheet_1.tables.keys())

# save file
wb.save("example/example.xlsx")

# close workbook to save memory
wb.close()