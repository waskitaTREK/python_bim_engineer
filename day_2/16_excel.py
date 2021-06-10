import openpyxl as xl

wb = xl.load_workbook("day_2/example.xlsx")

ws = wb['Sheet1']

# write
ws['F7'] = 123456

# read
print(ws['A2'])

wb.save("day_2/example.xlsx")

wb.close()